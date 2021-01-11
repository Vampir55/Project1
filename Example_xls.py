# Example works with xls tablets

import openpyxl

workbook = openpyxl.load_workbook(filename="Data1.xlsx")  # load xls file
sheet = workbook['Лист1']
sheet2 = workbook.create_sheet("List2")

for row in range(1, 20):
    sheet2.cell(row, 1, value=sheet.cell(row, 3).value)
    print(sheet2.cell(row, 1).value)
print("End lines")